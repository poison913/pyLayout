import os
import tkinter as tk
from tkinter import filedialog
import tkinter.messagebox
import matplotlib
matplotlib.use('TkAgg')
import numpy as np
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
#from matplotlib.backends.backend_tkagg import NavigationToolbar2TkAgg
from matplotlib.font_manager import FontProperties
from matplotlib.figure import Figure
import openpyxl
import math
import copy
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Alignment, Side, Border
from openpyxl.drawing import image
from openpyxl.chart import Series,LineChart, Reference

class layout:
    def mainlayout(self):
        self.win = tk.Tk()

        self.win.title('水位流量关系计算')
        self.win.geometry('905x620')
        self.win.resizable(width=False, height=False)
        self.win.configure(background='cornflowerblue')
        self.win.protocol("WM_DELETE_WINDOW", lambda: self.on_closing())
        # 定义第二个容器
        frame_right = tk.Frame(self.win, bg="WhiteSmoke", name="_right")
        # frame_right.place(relx=0.5, rely=0.0, relwidth=0.5, relheight=1)
        frame_right.place(x=455, y=0, height=620, width=445)

        self.rightlayout(frame_right)


        # 定义第一个容器
        frame_left = tk.Frame(self.win, height=620, width=445, bg="WhiteSmoke", name="_left")
        # frame_left.place(relx=0.0, rely=0, relwidth=0.5, relheight=1)
        frame_left.place(x=5, y=0, height=620, width=445)

        self.leftlauyout(frame_left)

        return self.win
    #左边frame
    def leftlauyout(self,frame_left):
        #第一行
        lable1 = tk.Label(frame_left,text = " 左边界:"  , bg="WhiteSmoke",font = ("宋体",13))
        lable1.place(x=7,y=2,height = 30, width=70)

        self.entry1 = tk.Entry(frame_left,borderwidth=2)
        self.entry1.place(x=80,y=2,height = 30, width=120)

        lable2 = tk.Label(frame_left,text = "  右边界:"  , bg="WhiteSmoke",font = ("宋体",13))
        lable2.place(x=220,y=2,height = 30, width=70)

        self.entry2 = tk.Entry(frame_left,borderwidth=2)
        self.entry2.place(x=300,y=2,height = 30, width=120)

        #第2行
        lable3 = tk.Label(frame_left,text = "   比降:"  , bg="WhiteSmoke",font = ("宋体",13))
        lable3.place(x=7,y=35,height = 30, width=70)

        self.entry3 = tk.Entry(frame_left,borderwidth=2)
        self.entry3.place(x=80,y=35,height = 30, width=120)

        lable4 = tk.Label(frame_left,text = "主槽糙率:"  , bg="WhiteSmoke",font = ("宋体",13))
        lable4.place(x=220,y=35,height = 30, width=70)

        self.entry4 = tk.Entry(frame_left,borderwidth=2)
        self.entry4.place(x=300,y=35,height = 30, width=120)


        #第3行
        lable5 = tk.Label(frame_left,text = "   水位:"  , bg="WhiteSmoke",font = ("宋体",13))
        lable5.place(x=7,y=68,height = 30, width=70)

        self.entry5 = tk.Entry(frame_left,borderwidth=2)
        self.entry5.place(x=80,y=68,height = 30, width=120)

        lable6 = tk.Label(frame_left,text = "    备注:"  , bg="WhiteSmoke",font = ("宋体",13))
        lable6.place(x=220,y=68,height = 30, width=70)

        self.entry6 = tk.Entry(frame_left,borderwidth=2)
        self.entry6.place(x=300,y=68,height = 30, width=120)
        #第4行
        lable7 = tk.Label(frame_left,text = "左滩边界:"  , bg="WhiteSmoke",font = ("宋体",13))
        lable7.place(x=7,y=101,height = 30, width=70)

        self.entry7 = tk.Entry(frame_left,borderwidth=2)
        self.entry7.place(x=80,y=101,height = 30, width=120)

        lable8 = tk.Label(frame_left,text = "右滩边界:"  , bg="WhiteSmoke",font = ("宋体",13))
        lable8.place(x=220,y=101,height = 30, width=70)

        self.entry8 = tk.Entry(frame_left,borderwidth=2)
        self.entry8.place(x=300,y=101,height = 30, width=120)

        #第5行
        lable9 = tk.Label(frame_left,text = "左滩糙率:"  , bg="WhiteSmoke",font = ("宋体",13))
        lable9.place(x=7,y=134,height = 30, width=70)

        self.entry9 = tk.Entry(frame_left,borderwidth=2)
        self.entry9.place(x=80,y=134,height = 30, width=120)

        lable10 = tk.Label(frame_left,text = "右滩糙率:"  , bg="WhiteSmoke",font = ("宋体",13))
        lable10.place(x=220,y=134,height = 30, width=70)

        self.entry10 = tk.Entry(frame_left,borderwidth=2)
        self.entry10.place(x=300,y=134,height = 30, width=120)
        #操作列
        frame_oprate = tk.Frame(frame_left,   bg="LightCyan")
        frame_oprate.place(x=0,y=167,height = 40, width=445)

        self.button1 = tk.Button(frame_left, text = "计算",background='DodgerBlue',command = lambda: self.calButton())
        self.button1.place(x=120,y=172,height = 30, width=50)

        self.button2 = tk.Button(frame_left, text = "计算水位流量关系",background='DodgerBlue',
                                 state ='disabled', command = lambda: self.showChart())
        self.button2.place(x=190,y=172,height = 30, width=150)

        #左槽
        group1 = tk.LabelFrame(frame_left, text="左滩", padx=5, pady=5)
        group1.place(x=10,y=205,height = 100, width=425)

        lable11 = tk.Label(group1,text = "左滩湿周:"  , bg="WhiteSmoke",font = ("宋体",13))
        lable11.place(x=7,y=3,height = 30, width=70)

        self.entry11 = tk.Entry(group1,borderwidth=2)
        self.entry11.place(x=80,y=3,height = 30, width=100)

        lable12 = tk.Label(group1,text = "左滩断面面积:"  , bg="WhiteSmoke",font = ("宋体",13))
        lable12.place(x=185,y=3,height = 30, width=120)

        self.entry12 = tk.Entry(group1,borderwidth=2)
        self.entry12.place(x=300,y=3,height = 30, width=105)


        lable13 = tk.Label(group1,text = "左滩水力半径:"  , bg="WhiteSmoke",font = ("宋体",13))
        lable13.place(x=7,y=39,height = 30, width=110)

        self.entry13 = tk.Entry(group1,borderwidth=2)
        self.entry13.place(x=120,y=39,height = 30, width=100)

        lable14 = tk.Label(group1,text = "左滩流量:"  , bg="WhiteSmoke",font = ("宋体",13))
        lable14.place(x=225,y=39,height = 30, width=80)

        self.entry14 = tk.Entry(group1,borderwidth=2)
        self.entry14.place(x=305,y=39,height = 30, width=100)

        # 右槽
        group2 = tk.LabelFrame(frame_left, text="右滩", padx=5, pady=5)
        group2.place(x=10, y=308, height=100, width=425)

        lable15 = tk.Label(group2, text="右滩湿周:", bg="WhiteSmoke", font=("宋体", 13))
        lable15.place(x=7, y=3, height=30, width=70)

        self.entry15 = tk.Entry(group2, borderwidth=2)
        self.entry15.place(x=80, y=3, height=30, width=100)

        lable16 = tk.Label(group2, text="右滩断面面积:", bg="WhiteSmoke", font=("宋体", 13))
        lable16.place(x=185, y=3, height=30, width=120)

        self.entry16 = tk.Entry(group2, borderwidth=2)
        self.entry16.place(x=300, y=3, height=30, width=105)

        lable17 = tk.Label(group2, text="右滩水力半径:", bg="WhiteSmoke", font=("宋体", 13))
        lable17.place(x=7, y=39, height=30, width=110)

        self.entry17 = tk.Entry(group2, borderwidth=2)
        self.entry17.place(x=120, y=39, height=30, width=100)

        lable18 = tk.Label(group2, text="右滩流量:", bg="WhiteSmoke", font=("宋体", 13))
        lable18.place(x=225, y=39, height=30, width=80)

        self.entry18 = tk.Entry(group2, borderwidth=2)
        self.entry18.place(x=305, y=39, height=30, width=100)

        # 主槽
        group3 = tk.LabelFrame(frame_left, text="主槽", padx=5, pady=5)
        group3.place(x=10, y=411, height=100, width=425)

        lable19 = tk.Label(group3, text="主槽湿周:", bg="WhiteSmoke", font=("宋体", 13))
        lable19.place(x=7, y=3, height=30, width=70)

        self.entry19 = tk.Entry(group3, borderwidth=2)
        self.entry19.place(x=80, y=3, height=30, width=100)

        lable20 = tk.Label(group3, text="主槽断面面积:", bg="WhiteSmoke", font=("宋体", 13))
        lable20.place(x=185, y=3, height=30, width=120)

        self.entry20 = tk.Entry(group3, borderwidth=2)
        self.entry20.place(x=300, y=3, height=30, width=105)

        lable21 = tk.Label(group3, text="主槽水力半径:", bg="WhiteSmoke", font=("宋体", 13))
        lable21.place(x=7, y=39, height=30, width=110)

        self.entry21 = tk.Entry(group3, borderwidth=2)
        self.entry21.place(x=120, y=39, height=30, width=100)

        lable22 = tk.Label(group3, text="主槽流量:", bg="WhiteSmoke", font=("宋体", 13))
        lable22.place(x=225, y=39, height=30, width=80)

        self.entry22 = tk.Entry(group3, borderwidth=2)
        self.entry22.place(x=305, y=39, height=30, width=100)

        # 总计
        group4 = tk.LabelFrame(frame_left, text="总计", padx=5, pady=5)
        group4.place(x=10, y=514, height=100, width=425)

        lable23 = tk.Label(group4, text="湿周:", bg="WhiteSmoke", font=("宋体", 13))
        lable23.place(x=7, y=3, height=30, width=70)

        self.entry23 = tk.Entry(group4, borderwidth=2)
        self.entry23.place(x=80, y=3, height=30, width=100)

        lable24 = tk.Label(group4, text="过水断面面积:", bg="WhiteSmoke", font=("宋体", 13))
        lable24.place(x=185, y=3, height=30, width=120)

        self.entry24 = tk.Entry(group4, borderwidth=2)
        self.entry24.place(x=300, y=3, height=30, width=105)

        lable25 = tk.Label(group4, text="水力半径:", bg="WhiteSmoke", font=("宋体", 13))
        lable25.place(x=7, y=39, height=30, width=110)

        self.entry25 = tk.Entry(group4, borderwidth=2)
        self.entry25.place(x=120, y=39, height=30, width=100)

        lable26 = tk.Label(group4, text="流量:", bg="WhiteSmoke", font=("宋体", 13))
        lable26.place(x=225, y=39, height=30, width=80)

        self.entry26 = tk.Entry(group4, borderwidth=2)
        self.entry26.place(x=305, y=39, height=30, width=100)

        #备注
        # lable27 = tk.Label(frame_left,text=" 备注:", bg="WhiteSmoke",font = ("宋体",13))
        # lable27.place(x=7,y=584,height = 30, width=70)
        #
        # self.entry27 = tk.Entry(frame_left,borderwidth=2)
        # self.entry27.place(x=80,y=584,height = 30, width=350)
    #右边frame
    def rightlayout(self,frame_right):

        #canvas
        self.canvas1 = tk.Frame(frame_right, bg="WhiteSmoke", name="_right")
        self.canvas1.place(x=0, y=40, height=580, width=445)
        fig = Figure(figsize=(5, 4), dpi=100, facecolor='WhiteSmoke')
        self.canvas1.ax = fig.add_subplot(111)
        self.canvas1.canvas = FigureCanvasTkAgg(fig, master=self.canvas1)
        self.canvas1.canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)
        self.canvas1.canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=1)
        # toolbar = NavigationToolbar2TkAgg(frame_right.canvas, frame_right)
        # toolbar.update()
        # 操作列
        button1 = tk.Button(frame_right,
                            text="选择断面数据",
                            command=lambda: self.openfile(frame_right)
                            )
        button1.place(x=10, y=2, height=30, width=120)
        #self.draw(canvas1)
    #画图
    def draw(self):
        #'''绘图逻辑'''
        font_set = FontProperties(fname=r"c:\windows\fonts\simsun.ttc", size=12)

        # x = range(2, 26, 2)
        # y = [15, 13, 14.5, 17, 20, 25, 26, 26, 24, 22, 18, 15]
        # y1 = [17, 17, 17, 17, 17,17, 17, 17, 17, 17, 17, 17]
        # 数据在y周的位置是一个可迭代的对象
        # x轴 y轴的数据一起组成了所有要绘制出的图标
        # self.fig.clf() # 方式一：①清除整个Figure区域
        # self.ax = self.fig.add_subplot(111) # ②重新分配Axes区域
        self.canvas1.ax.clear() # 方式二：①清除原来的Axes区域
        # 添加描述信息
        self.canvas1.ax.set_xlabel('起点距', fontproperties=font_set )
        self.canvas1.ax.set_ylabel('高程', fontproperties=font_set )
        self.canvas1.ax.plot(self.x, self.y,linewidth=3,color='blue',marker='.',markerfacecolor='red',markersize=10)  # 传入x和y 通过plot绘制出折线图
        #canvas1.ax.plot(self.x, self.y1,linewidth=3,color='red')  # 传入x和y 通过plot绘制出折线图
        if (self.is_number(self.entry5.get()) == True):
            self.canvas1.ax.axhline(y=float(self.entry5.get()), ls="-", c="green")#水位
        if (self.is_number(self.entry7.get()) == True):
            self.canvas1.ax.axvline(x=float(self.entry7.get()), ls="-", c="green") #左滩边界
        if (self.is_number(self.entry8.get()) == True):
            self.canvas1.ax.axvline(x=float(self.entry8.get()), ls="-", c="green")#右滩边界
        self.canvas1.canvas.draw()
    #读取excel
    def openfile(self,frame_right ):
        # if(self.is_number(self.entry5.get())== False):
        #     tk.messagebox.showerror(title='错误', message="水位数据不能为空，且必须为数字")
        #     return
        my_filetypes = [('text files', '.xlsx')]#('all files', '.*'), ('text files', '.xls'),
        answer = filedialog.askopenfilename(parent=frame_right,
                                            initialdir=os.getcwd(),
                                            title="Please select a file:",
                                            filetypes=my_filetypes)
        try:
            wb = openpyxl.load_workbook(answer)
            ws1 = wb.get_sheet_by_name("断面")
            #ws2 = wb.get_sheet_by_name("水位")
            #self.waterlevel = ws2.cell(row=1, column=1).value
            #self.waterlevel = float(self.entry5.get())
            # print(ws1.cell(row=1, column=2).value)
            self.x = []
            self.y = []
            self.y1 = []
            self.basexy = []
            for num in range(2, ws1._current_row):
                self.x.append(ws1.cell(row=num, column=1).value)
                self.y.append(ws1.cell(row=num, column=2).value)
                #self.y1.append(self.waterlevel)
                self.basexy.append([ws1.cell(row=num, column=1).value,ws1.cell(row=num, column=2).value])
            #插入默认值
            self.entry1.insert(0, str(self.x[0]))
            self.entry2.insert(0, str(self.x[len(self.x)-1]))
            self.draw()
        except Exception as e:
            #print(e)  # 打印所有异常到屏幕
            tk.messagebox.showerror(title='错误', message="excel文档错误："+e)
    #计算流量
    def calarea(self):
        self.xy = copy.deepcopy(self.basexy)
        self.waterlevel = float(self.entry5.get()) # 水位
        #获取所有交点
        allpoint=[];
        for i in range(len(self.y)):
            if(i>0 and (self.y[i]-self.waterlevel)*(self.y[i-1]-self.waterlevel)<=0):
                # 一般式 Ax+By+C=0
                a = self.y[i] - self.y[i-1]
                b = self.x[i-1] - self.x[i]
                c = self.x[i] * self.y[i-1] - self.x[i-1] * self.y[i]
                xpoint = ((0-b*self.waterlevel)-c)/a
                allpoint.append([xpoint,self.waterlevel])
        if(len(allpoint)==0):
            tk.messagebox.showerror(title='错误', message="交点少于一个，无法计算面积")
            return
        rightpoint = []#右交点
        leftpoint = []#左交点
        if(len(allpoint)==1):#仅有一个交点的情况下
            if(allpoint[0][0]- float(self.entry1.get()) > allpoint[0][0]- float(self.entry2.get())):#唯一的焦点距离左边界比距离右边界远   则左交点为左边界  唯一的交点为右交点
                rightpoint = allpoint[0]
                for i in range(len(self.y)):#求左边界对应的y值
                    if (i > 0 and  (self.x[i]-float(self.entry1.get()) )*(self.x[i-1]-float(self.entry1.get()) )<= 0):
                        # 一般式 Ax+By+C=0
                        a = self.y[i] - self.y[i - 1]
                        b = self.x[i - 1] - self.x[i]
                        c = self.x[i] * self.y[i - 1] - self.x[i - 1] * self.y[i]
                        ypoint = ((0 - a * float(self.entry1.get())) - c) / b
                        leftpoint = [float(self.entry1.get()),ypoint]
            else:
                leftpoint = allpoint[0]
                for i in range(len(self.y)):#求右边界对应的y值
                    if (i > 0 and  (self.x[i]-float(self.entry2.get()) )*(self.x[i-1]-float(self.entry2.get()) )<= 0):
                        # 一般式 Ax+By+C=0
                        a = self.y[i] - self.y[i - 1]
                        b = self.x[i - 1] - self.x[i]
                        c = self.x[i] * self.y[i - 1] - self.x[i - 1] * self.y[i]
                        ypoint = ((0 - a * float(self.entry2.get())) - c) / b
                        rightpoint = [float(self.entry2.get()),ypoint]
        else:
            # 获取相聚最远的两个相邻焦点
            index = 0
            maxlengh = 0;
            for i in range(len(allpoint)):
                if (i > 0 and allpoint[i][0] - allpoint[i - 1][0] > maxlengh):
                    index = i
                    maxlengh = allpoint[i][0] - allpoint[i - 1][0]
            rightpoint = allpoint[index]  # 右交点
            leftpoint = allpoint[index - 1]  # 左交点
        #将焦点加入xy数据
        self.xy.append(leftpoint)
        self.xy.append(rightpoint)
        self.xy.sort()
        #求左中右三块的面积
        #左滩边界
        leftlength = float(self.entry7.get())
        #右滩边界
        rightlength = float(self.entry8.get())

        self.area = 0#总面积
        self.leftarea = 0 #左滩面积
        self.rightarea = 0 #右滩面积
        self.midarea = 0 #中间面积

        if(leftlength<0 or leftlength> self.x[len(self.x)-1]):
            tk.messagebox.showerror(title='左滩边界数据错误',message="左滩边界应大于0，小于"+str(self.x[len(self.x)-1]))
            return
        if (rightlength < 0 or rightlength > self.x[len(self.x)-1]):
            tk.messagebox.showerror(title='右滩边界数据错误', message="右滩边界应大于0，小于" + str(self.x[len(self.x)-1]))
            return
        if(leftlength>= rightlength):
            tk.messagebox.showerror(title='数据错误', message="左滩边界应小于右滩边界")
            return
        #计算左右滩边界点位
        for i in range(len(self.x)):
            if(i>0 and (self.x[i]-leftlength)*(self.x[i-1]-leftlength)<=0):
                # 一般式 Ax+By+C=0
                a = self.y[i] - self.y[i-1]
                b = self.x[i-1] - self.x[i]
                c = self.x[i] * self.y[i-1] - self.x[i-1] * self.y[i]
                ypoint = ((0-a*leftlength)-c)/b
                self.xy.append([leftlength,ypoint])
            if (i > 0 and (self.x[i] - rightlength) * (self.x[i - 1] - rightlength) <= 0):
                # 一般式 Ax+By+C=0
                a = self.y[i] - self.y[i - 1]
                b = self.x[i - 1] - self.x[i]
                c = self.x[i] * self.y[i - 1] - self.x[i - 1] * self.y[i]
                ypoint = ((0 - a * rightlength) - c) / b
                self.xy.append([rightlength, ypoint])
        self.xy.sort()
        mainleft = 0 #主槽左侧
        mainright= 0 #主槽右侧距离
        if (leftlength <= leftpoint[0]):
            mainleft = leftpoint[0]
        else:
            mainleft = leftlength
        if (rightlength >= rightpoint[0]):
            mainright = rightpoint[0]
        else:
            mainright = rightlength
        #算面积
        for i  in range(len(self.xy)):
            if(self.xy[i][0]>leftpoint[0] and self.xy[i][0] <= rightpoint[0] ):
                self.area += (self.waterlevel- self.xy[i][1] +self.waterlevel- self.xy[i-1][1] ) *(self.xy[i][0]-self.xy[i-1][0])/2   #总面积  采用梯形面积计算方法
            if (self.xy[i][0]>leftpoint[0] and self.xy[i][0] <= mainleft):
                self.leftarea += (self.waterlevel- self.xy[i][1] +self.waterlevel- self.xy[i-1][1] ) *(self.xy[i][0]-self.xy[i-1][0])/2  #左面积
            elif (self.xy[i][0] > mainleft and self.xy[i][0] <= mainright):
                self.midarea += (self.waterlevel- self.xy[i][1] +self.waterlevel- self.xy[i-1][1] ) *(self.xy[i][0]-self.xy[i-1][0])/2   # 主面积
            elif (self.xy[i][0] > mainright and  self.xy[i][0] <= rightpoint[0]):
                self.rightarea += (self.waterlevel- self.xy[i][1] +self.waterlevel- self.xy[i-1][1] ) *(self.xy[i][0]-self.xy[i-1][0])/2   # 右面积

        #tk.messagebox.showinfo(title='面积结果', message="总面积："+ str(area) +"。左槽面积："+str(leftarea) +"。主槽面积："+str(mainarea)+"。右槽面积："+str(rightarea))
        #计算水力半径  --先算湿周
        self.leftWetCycle = 0;
        self.rightWetCycle = 0;
        self.midWetCycle = 0;
        self.leftWaterRadiu = 0 #左滩水力半径
        self.rightWaterRadiu = 0 #右滩水力半径
        self.midWaterRadiu = 0 #主水力半径
        for i  in range(len(self.xy)):
            if (self.xy[i][0] > leftpoint[0] and self.xy[i][0] <= mainleft):
                self.leftWetCycle += math.hypot(self.xy[i][0]-self.xy[i-1][0],self.xy[i][1]-self.xy[i-1][1])
            elif (self.xy[i][0] <= mainright):
                self.midWetCycle += math.hypot(self.xy[i][0]-self.xy[i-1][0],self.xy[i][1]-self.xy[i-1][1])
            elif(self.xy[i][0] <= rightpoint[0]):
                self.rightWetCycle += math.hypot(self.xy[i][0]-self.xy[i-1][0],self.xy[i][1]-self.xy[i-1][1])
            if (self.xy[i][0] == mainleft or self.xy[i][0] == mainright):
                self.midWetCycle += (self.waterlevel - self.xy[i][1])
        #水力半径=面积/湿周
        self.leftWaterRadiu = 0 if self.leftWetCycle == 0 else self.leftarea/self.leftWetCycle
        self.rightWaterRadiu = 0 if self.rightWetCycle == 0 else self.rightarea/self.rightWetCycle
        self.midWaterRadiu = 0 if self.midWetCycle == 0 else self.midarea/self.midWetCycle
        #计算流量
        self.descendingRate = float(self.entry3.get()) #比降
        self.leftFlow = (self.descendingRate**(1/2))  * self.leftarea * (self.leftWaterRadiu**(2/3))/float(self.entry9.get())
        self.rightFlow =(self.descendingRate**(1/2))  * self.rightarea * (self.rightWaterRadiu**(2/3))/float(self.entry10.get())
        self.midFlow =(self.descendingRate**(1/2))  * self.midarea * (self.midWaterRadiu**(2/3))/float(self.entry4.get())
        #统计数据
        self.allWaterRadiu = self.leftWaterRadiu +self.rightWaterRadiu +self.midWaterRadiu  #水力半径
        self.allWetCycle = self.leftWetCycle +self.rightWetCycle +self.midWetCycle  #湿周
        self.allArea = self.area  #面积
        self.allFlow = self.leftFlow +self.rightFlow +self.midFlow  #湿周

        #数据写入界面 需要清空
        self.entry11.delete(0,99999)
        self.entry12.delete(0,99999)
        self.entry13.delete(0,99999)
        self.entry14.delete(0,99999)
        self.entry15.delete(0,99999)
        self.entry16.delete(0,99999)
        self.entry17.delete(0,99999)
        self.entry18.delete(0,99999)
        self.entry19.delete(0,99999)
        self.entry20.delete(0,99999)
        self.entry21.delete(0,99999)
        self.entry22.delete(0,99999)
        self.entry23.delete(0,99999)
        self.entry24.delete(0,99999)
        self.entry25.delete(0,99999)
        self.entry26.delete(0,99999)
        self.entry11.insert(0, str(self.leftWetCycle))
        self.entry12.insert(0, str(self.leftarea))
        self.entry13.insert(0, str(self.leftWaterRadiu))
        self.entry14.insert(0, str(self.leftFlow))

        self.entry15.insert(0, str(self.rightWetCycle))
        self.entry16.insert(0, str(self.rightarea))
        self.entry17.insert(0, str(self.rightWaterRadiu))
        self.entry18.insert(0, str(self.rightFlow))

        self.entry19.insert(0, str(self.midWetCycle))
        self.entry20.insert(0, str(self.midarea))
        self.entry21.insert(0, str(self.midWaterRadiu))
        self.entry22.insert(0, str(self.midFlow))

        self.entry23.insert(0, str(self.allWetCycle))
        self.entry24.insert(0, str(self.allArea))
        self.entry25.insert(0, str(self.allWaterRadiu))
        self.entry26.insert(0, str(self.allFlow))

    #判断输入的是数字
    def is_number(self,s):
        try:
            float(s)
            return True
        except ValueError:
            pass
        return False
    #计算按钮点击事件
    def calButton(self):
        if(hasattr(self,'x')==False):
            tk.messagebox.showinfo(title='提示', message="请先打开断面数据！")
            return
        if (self.is_number(self.entry1.get()) == False):
            tk.messagebox.showerror(title='错误', message="左边界不能为空，且必须为数字")
            return
        if (self.is_number(self.entry2.get()) == False):
            tk.messagebox.showerror(title='错误', message="右边界不能为空，且必须为数字")
            return
        if(self.is_number(self.entry3.get())== False):
            tk.messagebox.showerror(title='错误', message="比降不能为空，且必须为数字")
            return
        if (self.is_number(self.entry4.get()) == False):
            tk.messagebox.showerror(title='错误', message="主槽糙率不能为空，且必须为数字")
            return
        if (self.is_number(self.entry5.get()) == False):
            tk.messagebox.showerror(title='错误', message="水位不能为空，且必须为数字")
            return
        if (self.is_number(self.entry7.get()) == False):
            tk.messagebox.showerror(title='错误', message="左滩边界不能为空，且必须为数字")
            return
        if (self.is_number(self.entry8.get()) == False):
            tk.messagebox.showerror(title='错误', message="右滩边界不能为空，且必须为数字")
            return
        if(float(self.entry1.get())>float(self.entry7.get())  ):
            tk.messagebox.showerror(title='错误', message="左边界不能大于左滩边界")
            return
        if(float(self.entry2.get())<float(self.entry8.get())):
            tk.messagebox.showerror(title='错误', message="右边界不能小于右滩边界")
            return
        if (self.is_number(self.entry9.get()) == False):
            tk.messagebox.showerror(title='错误', message="左滩糙率不能为空，且必须为数字")
            return
        if (self.is_number(self.entry10.get()) == False):
            tk.messagebox.showerror(title='错误', message="右滩糙率不能为空，且必须为数字")
            return
        self.draw()
        self.calarea()
        #设置水位流量关系button可用
        self.button2['state']= 'normal'

    def on_closing(self):
        if (hasattr(self, 'chart') == True):
            if(len(self.chart.children)>0):
                self.chart.destroy()
        self.win.destroy()
    #水位流量关系按钮点击事件
    def showChart(self):
        self.chart = tk.Tk()
        self.chart.title('水位流量关系计算')
        self.chart.geometry('675x520')
        self.chart.resizable(width=False, height=False)
        self.chart.configure(background='cornflowerblue')
        self.chart.wm_attributes('-topmost', 1)
        # 定义第一个容器
        chart_left = tk.Frame(self.chart, height=520, width=445, bg="WhiteSmoke")
        # frame_left.place(relx=0.0, rely=0, relwidth=0.5, relheight=1)
        chart_left.place(x=5, y=0, height=520, width=445)
        # canvas
        self.canvas2 = tk.Frame(chart_left, bg="WhiteSmoke", name="_right")
        self.canvas2.place(x=0, y=0, height=500, width=445)
        fig = Figure(figsize=(5, 4), dpi=100, facecolor='WhiteSmoke')
        self.canvas2.ax = fig.add_subplot(111)
        self.canvas2.canvas = FigureCanvasTkAgg(fig, master=self.canvas2)
        self.canvas2.canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)
        self.canvas2.canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=1)
        # 定义第2个容器
        chart_right = tk.Frame(self.chart, height=520, width=21, bg="WhiteSmoke")
        # frame_left.place(relx=0.0, rely=0, relwidth=0.5, relheight=1)
        chart_right.place(x=455, y=0, height=520, width=215)
        lable1 = tk.Label(chart_right, text=" 左边界:", bg="WhiteSmoke", font=("宋体", 13))
        lable1.place(x=7, y=8, height=30, width=70)

        self.chartEnter1 = tk.Entry(chart_right, borderwidth=2)
        self.chartEnter1.place(x=80, y=8, height=30, width=120)

        lable2 = tk.Label(chart_right, text=" 右边界:", bg="WhiteSmoke", font=("宋体", 13))
        lable2.place(x=7, y=46, height=30, width=70)

        self.chartEnter2 = tk.Entry(chart_right, borderwidth=2)
        self.chartEnter2.place(x=80, y=46, height=30, width=120)

        lable3 = tk.Label(chart_right, text=" 比 降:", bg="WhiteSmoke", font=("宋体", 13))
        lable3.place(x=7, y=84, height=30, width=70)

        self.chartEnter3 = tk.Entry(chart_right, borderwidth=2)
        self.chartEnter3.place(x=80, y=84, height=30, width=120)

        lable4 = tk.Label(chart_right, text="主槽糙率:", bg="WhiteSmoke", font=("宋体", 13))
        lable4.place(x=7, y=122, height=30, width=70)

        self.chartEnter4 = tk.Entry(chart_right, borderwidth=2)
        self.chartEnter4.place(x=80, y=122, height=30, width=120)

        lable5 = tk.Label(chart_right, text="左滩边界:", bg="WhiteSmoke", font=("宋体", 13))
        lable5.place(x=7, y=160, height=30, width=70)

        self.chartEnter5 = tk.Entry(chart_right, borderwidth=2)
        self.chartEnter5.place(x=80, y=160, height=30, width=120)

        lable6 = tk.Label(chart_right, text="右滩边界:", bg="WhiteSmoke", font=("宋体", 13))
        lable6.place(x=7, y=198, height=30, width=70)

        self.chartEnter6 = tk.Entry(chart_right, borderwidth=2)
        self.chartEnter6.place(x=80, y=198, height=30, width=120)

        lable7 = tk.Label(chart_right, text="左滩糙率:", bg="WhiteSmoke", font=("宋体", 13))
        lable7.place(x=7, y=236, height=30, width=70)

        self.chartEnter7 = tk.Entry(chart_right, borderwidth=2)
        self.chartEnter7.place(x=80, y=236, height=30, width=120)

        lable8 = tk.Label(chart_right, text="右滩糙率:", bg="WhiteSmoke", font=("宋体", 13))
        lable8.place(x=7, y=274, height=30, width=70)

        self.chartEnter8 = tk.Entry(chart_right, borderwidth=2)
        self.chartEnter8.place(x=80, y=274, height=30, width=120)

        self.chartbutton1 = tk.Button(chart_right, text="调整", background='DodgerBlue' ,command = lambda: self.calchart())
        self.chartbutton1.place(x=7, y=374, height=30, width=50)

        self.chartbutton2 = tk.Button(chart_right, text="导出水位流量关系", background='DodgerBlue'  ,command = lambda: self.toexcel())
        self.chartbutton2.place(x=80, y=374, height=30, width=120)
        #设置弹出框参数
        self.setChartValue()
        #计算水位流量关系
        self.calchart()
        self.chart.mainloop()
    #设置弹出框参数
    def setChartValue(self):
        self.chartEnter1.delete(0,99999)
        self.chartEnter2.delete(0,99999)
        self.chartEnter3.delete(0,99999)
        self.chartEnter4.delete(0,99999)
        self.chartEnter5.delete(0,99999)
        self.chartEnter6.delete(0,99999)
        self.chartEnter7.delete(0,99999)
        self.chartEnter8.delete(0,99999)

        self.chartEnter1.insert(0, self.entry1.get())
        self.chartEnter2.insert(0, self.entry2.get())
        self.chartEnter3.insert(0, self.entry3.get())
        self.chartEnter4.insert(0, self.entry4.get())
        self.chartEnter5.insert(0, self.entry7.get())
        self.chartEnter6.insert(0, self.entry8.get())
        self.chartEnter7.insert(0, self.entry9.get())
        self.chartEnter8.insert(0, self.entry10.get())

    # 计算水位流量关系
    def calchart(self):
        #获取水位最低点和左右两侧的最高点
        waterlevelmin = 999999999999999999999
        minIndex = 0;
        waterlevelleftMax = 0
        waterlevelrightMax = 0
        waterlevelmax = 0
        for i in range(len(self.xy)):
            if(self.xy[i][1]< waterlevelmin):
                waterlevelmin = self.xy[i][1]
                minIndex = i
        for i in range(len(self.xy)):
            if(i < minIndex):
                if(self.xy[i][1]>waterlevelleftMax):
                    waterlevelleftMax = self.xy[i][1]
            else:
                if (self.xy[i][1] > waterlevelrightMax):
                    waterlevelrightMax = self.xy[i][1]
        waterlevelmax = waterlevelrightMax if(waterlevelrightMax < waterlevelleftMax) else waterlevelleftMax
        self.levelFlow = [['水位','流量']];
        self.level=[];
        self.Flow = [];
        for i in np.arange(waterlevelmin, waterlevelmax, 0.1):
            #print(round(i, 1))
            currentFlow = self.calchartSingle(round(i, 1))
            self.level.append(i)
            self.Flow.append(currentFlow)
            self.levelFlow.append([i,currentFlow])
        self.drawChart2()
    def calchartSingle(self,waterlevel):
        self.xy = copy.deepcopy(self.basexy)
        self.waterlevel = waterlevel  # 水位
        # 获取所有交点
        allpoint = [];
        for i in range(len(self.y)):
            if (i > 0 and (self.y[i] - self.waterlevel) * (self.y[i - 1] - self.waterlevel) <= 0):
                # 一般式 Ax+By+C=0
                a = self.y[i] - self.y[i - 1]
                b = self.x[i - 1] - self.x[i]
                c = self.x[i] * self.y[i - 1] - self.x[i - 1] * self.y[i]
                xpoint = ((0 - b * self.waterlevel) - c) / a
                allpoint.append([xpoint, self.waterlevel])
        if (len(allpoint) == 0):
            #tk.messagebox.showerror(title='错误', message="交点少于一个，无法计算面积")
            return
        rightpoint = []  # 右交点
        leftpoint = []  # 左交点
        if (len(allpoint) == 1):  # 仅有一个交点的情况下
            if (allpoint[0][0] - float(self.chartEnter1.get()) > allpoint[0][0] - float(
                    self.chartEnter2.get())):  # 唯一的焦点距离左边界比距离右边界远   则左交点为左边界  唯一的交点为右交点
                rightpoint = allpoint[0]
                for i in range(len(self.y)):  # 求左边界对应的y值
                    if (i > 0 and (self.x[i] - float(self.chartEnter1.get())) * (
                            self.x[i - 1] - float(self.chartEnter1.get())) <= 0):
                        # 一般式 Ax+By+C=0
                        a = self.y[i] - self.y[i - 1]
                        b = self.x[i - 1] - self.x[i]
                        c = self.x[i] * self.y[i - 1] - self.x[i - 1] * self.y[i]
                        ypoint = ((0 - a * float(self.chartEnter1.get())) - c) / b
                        leftpoint = [float(self.chartEnter1.get()), ypoint]
            else:
                leftpoint = allpoint[0]
                for i in range(len(self.y)):  # 求右边界对应的y值
                    if (i > 0 and (self.x[i] - float(self.chartEnter2.get())) * (
                            self.x[i - 1] - float(self.chartEnter2.get())) <= 0):
                        # 一般式 Ax+By+C=0
                        a = self.y[i] - self.y[i - 1]
                        b = self.x[i - 1] - self.x[i]
                        c = self.x[i] * self.y[i - 1] - self.x[i - 1] * self.y[i]
                        ypoint = ((0 - a * float(self.chartEnter2.get())) - c) / b
                        rightpoint = [float(self.chartEnter2.get()), ypoint]
        else:
            # 获取相聚最远的两个相邻焦点
            index = 0
            maxlengh = 0;
            for i in range(len(allpoint)):
                if (i > 0 and allpoint[i][0] - allpoint[i - 1][0] > maxlengh):
                    index = i
                    maxlengh = allpoint[i][0] - allpoint[i - 1][0]
            rightpoint = allpoint[index]  # 右交点
            leftpoint = allpoint[index - 1]  # 左交点
        # 将焦点加入xy数据
        self.xy.append(leftpoint)
        self.xy.append(rightpoint)
        self.xy.sort()
        # 求左中右三块的面积
        # 左滩边界
        leftlength = float(self.chartEnter5.get())
        # 右滩边界
        rightlength = float(self.chartEnter6.get())

        self.area = 0  # 总面积
        self.leftarea = 0  # 左滩面积
        self.rightarea = 0  # 右滩面积
        self.midarea = 0  # 中间面积

        if (leftlength < 0 or leftlength > self.x[len(self.x) - 1]):
            tk.messagebox.showerror(title='左滩边界数据错误', message="左滩边界应大于0，小于" + str(self.x[len(self.x) - 1]))
            return
        if (rightlength < 0 or rightlength > self.x[len(self.x) - 1]):
            tk.messagebox.showerror(title='右滩边界数据错误', message="右滩边界应大于0，小于" + str(self.x[len(self.x) - 1]))
            return
        if (leftlength >= rightlength):
            tk.messagebox.showerror(title='数据错误', message="左滩边界应小于右滩边界")
            return
        # 计算左右滩边界点位
        for i in range(len(self.x)):
            if (i > 0 and (self.x[i] - leftlength) * (self.x[i - 1] - leftlength) <= 0):
                # 一般式 Ax+By+C=0
                a = self.y[i] - self.y[i - 1]
                b = self.x[i - 1] - self.x[i]
                c = self.x[i] * self.y[i - 1] - self.x[i - 1] * self.y[i]
                ypoint = ((0 - a * leftlength) - c) / b
                self.xy.append([leftlength, ypoint])
            if (i > 0 and (self.x[i] - rightlength) * (self.x[i - 1] - rightlength) <= 0):
                # 一般式 Ax+By+C=0
                a = self.y[i] - self.y[i - 1]
                b = self.x[i - 1] - self.x[i]
                c = self.x[i] * self.y[i - 1] - self.x[i - 1] * self.y[i]
                ypoint = ((0 - a * rightlength) - c) / b
                self.xy.append([rightlength, ypoint])
        self.xy.sort()
        mainleft = 0  # 主槽左侧
        mainright = 0  # 主槽右侧距离
        if (leftlength <= leftpoint[0]):
            mainleft = leftpoint[0]
        else:
            mainleft = leftlength
        if (rightlength >= rightpoint[0]):
            mainright = rightpoint[0]
        else:
            mainright = rightlength
        # 算面积
        for i in range(len(self.xy)):
            if (self.xy[i][0] > leftpoint[0] and self.xy[i][0] <= rightpoint[0]):
                self.area += (self.waterlevel - self.xy[i][1] + self.waterlevel - self.xy[i - 1][1]) * (
                        self.xy[i][0] - self.xy[i - 1][0]) / 2  # 总面积  采用梯形面积计算方法
            if (self.xy[i][0] > leftpoint[0] and self.xy[i][0] <= mainleft):
                self.leftarea += (self.waterlevel - self.xy[i][1] + self.waterlevel - self.xy[i - 1][1]) * (
                        self.xy[i][0] - self.xy[i - 1][0]) / 2  # 左面积
            elif (self.xy[i][0] > mainleft and self.xy[i][0] <= mainright):
                self.midarea += (self.waterlevel - self.xy[i][1] + self.waterlevel - self.xy[i - 1][1]) * (
                        self.xy[i][0] - self.xy[i - 1][0]) / 2  # 主面积
            elif (self.xy[i][0] > mainright and self.xy[i][0] <= rightpoint[0]):
                self.rightarea += (self.waterlevel - self.xy[i][1] + self.waterlevel - self.xy[i - 1][1]) * (
                        self.xy[i][0] - self.xy[i - 1][0]) / 2  # 右面积

        # tk.messagebox.showinfo(title='面积结果', message="总面积："+ str(area) +"。左槽面积："+str(leftarea) +"。主槽面积："+str(mainarea)+"。右槽面积："+str(rightarea))
        # 计算水力半径  --先算湿周
        self.leftWetCycle = 0;
        self.rightWetCycle = 0;
        self.midWetCycle = 0;
        self.leftWaterRadiu = 0  # 左滩水力半径
        self.rightWaterRadiu = 0  # 右滩水力半径
        self.midWaterRadiu = 0  # 主水力半径
        for i in range(len(self.xy)):
            if (self.xy[i][0] > leftpoint[0] and self.xy[i][0] <= mainleft):
                self.leftWetCycle += math.hypot(self.xy[i][0] - self.xy[i - 1][0], self.xy[i][1] - self.xy[i - 1][1])
            elif (self.xy[i][0] <= mainright):
                self.midWetCycle += math.hypot(self.xy[i][0] - self.xy[i - 1][0], self.xy[i][1] - self.xy[i - 1][1])
            elif (self.xy[i][0] <= rightpoint[0]):
                self.rightWetCycle += math.hypot(self.xy[i][0] - self.xy[i - 1][0], self.xy[i][1] - self.xy[i - 1][1])
            if (self.xy[i][0] == mainleft or self.xy[i][0] == mainright):
                self.midWetCycle += (self.waterlevel - self.xy[i][1])
        # 水力半径=面积/湿周
        self.leftWaterRadiu = 0 if self.leftWetCycle == 0 else self.leftarea / self.leftWetCycle
        self.rightWaterRadiu = 0 if self.rightWetCycle == 0 else self.rightarea / self.rightWetCycle
        self.midWaterRadiu = 0 if self.midWetCycle == 0 else self.midarea / self.midWetCycle
        # 计算流量
        self.descendingRate = float(self.chartEnter3.get())  # 比降
        self.leftFlow = (self.descendingRate ** (1 / 2)) * self.leftarea * (self.leftWaterRadiu ** (2 / 3)) / float(
            self.chartEnter7.get())
        self.rightFlow = (self.descendingRate ** (1 / 2)) * self.rightarea * (self.rightWaterRadiu ** (2 / 3)) / float(
            self.chartEnter8.get())
        self.midFlow = (self.descendingRate ** (1 / 2)) * self.midarea * (self.midWaterRadiu ** (2 / 3)) / float(
            self.chartEnter4.get())
        return self.leftFlow + self.rightFlow + self.midFlow

    def drawChart2(self):
        # '''绘图逻辑'''
        font_set = FontProperties(fname=r"c:\windows\fonts\simsun.ttc", size=12)

        # x = range(2, 26, 2)
        # y = [15, 13, 14.5, 17, 20, 25, 26, 26, 24, 22, 18, 15]
        # y1 = [17, 17, 17, 17, 17,17, 17, 17, 17, 17, 17, 17]
        # 数据在y周的位置是一个可迭代的对象
        # x轴 y轴的数据一起组成了所有要绘制出的图标
        # self.fig.clf() # 方式一：①清除整个Figure区域
        # self.ax = self.fig.add_subplot(111) # ②重新分配Axes区域
        self.canvas2.ax.clear()  # 方式二：①清除原来的Axes区域
        # 添加描述信息
        self.canvas2.ax.set_xlabel('流量(m³/s)', fontproperties=font_set)
        self.canvas2.ax.set_ylabel('水位(m)', fontproperties=font_set)
        self.canvas2.ax.plot(self.Flow, self.level, linewidth=3, color='blue', marker='.', markerfacecolor='red',
                             markersize=10)  # 传入x和y 通过plot绘制出折线图
        self.canvas2.canvas.draw()
    #导出excel
    def toexcel(self):
        wb = Workbook()
        ws = wb.active

        for row in self.levelFlow:
            ws.append(row)

        # chart = LineChart()  # 图表对象
        # data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=len(self.levelFlow))  # 涉及数据
        # seriesObj = Series(data, title='水位流量关系')  # 创建series对象
        # chart.y_axis.title = "水位"
        # chart.x_axis.title = "流量"
        # chart.append(seriesObj)  # 添加到chart中
        # ws.add_chart(chart, "A6")  # 将图表添加到 sheet中
        wb.save("d:\\sample.xlsx")
win = layout()
win.mainlayout().mainloop()