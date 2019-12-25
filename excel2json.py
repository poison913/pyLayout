import openpyxl
import os
import tkinter as tk
from tkinter import filedialog
import json
# encoding: utf-8
def openfile( ):
    my_filetypes = [('text files', '.xlsx')]  # ('all files', '.*'), ('text files', '.xls'),
    answer = filedialog.askopenfilename(
                                        initialdir=os.getcwd(),
                                        title="Please select a file:",
                                        filetypes=my_filetypes)
    try:
        wb = openpyxl.load_workbook(answer)
        ws1 = wb.get_active_sheet()
        title=[]
        list = []
        for num in range(1, ws1.max_column+1):
            title.append(ws1.cell(row=1, column=num).value)
        for num in range(2, ws1._current_row+1):
            info = dict()
            for colNum in range(0, len(title)):
                info[title[colNum]] = ws1.cell(row=num, column=colNum+1).value
            list.append(info)
        json_data = json.dumps(list,ensure_ascii=False)
        #print(unicode("json_data", encoding="utf-8"))
        print(json_data)
    except Exception as e:
        # print(e)  # 打印所有异常到屏幕
        tk.messagebox.showerror(title='错误', message="excel文档错误：" + e)

openfile()